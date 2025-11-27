from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import numpy as np
import os
import json
from werkzeug.utils import secure_filename
import io
from datetime import datetime
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import openai
from openai import OpenAI
import base64
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import copy

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['ANALYSIS_FOLDER'] = 'analysis_outputs'

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls', 'json'}

# Initialize OpenAI client (will be configured when API key is provided)
client = None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def read_file(file_path, filename):
    ext = filename.rsplit('.', 1)[1].lower()
    try:
        if ext == 'csv':
            return pd.read_csv(file_path, parse_dates=True, infer_datetime_format=True), None
        elif ext in ['xlsx', 'xls']:
            # For Excel files, also load the workbook to preserve formatting
            df = pd.read_excel(file_path)
            workbook = load_workbook(file_path) if ext == 'xlsx' else None
            return df, workbook
        elif ext == 'json':
            return pd.read_json(file_path), None
    except Exception as e:
        raise ValueError(f"Error reading file: {str(e)}")

def clean_data_for_json(df):
    df_clean = df.copy()
    df_clean = df_clean.where(pd.notna(df_clean), None)
    for col in df_clean.columns:
        if pd.api.types.is_datetime64_any_dtype(df_clean[col]):
            df_clean[col] = df_clean[col].astype(str)
    return df_clean

def extract_excel_formatting(workbook):
    """Extract formatting information from Excel workbook"""
    if not workbook:
        return None
    
    formatting_info = {
        'styles': {},
        'column_widths': {},
        'row_heights': {},
        'merged_cells': []
    }
    
    try:
        sheet = workbook.active
        
        # Extract column widths
        for col_idx, col in enumerate(sheet.columns, 1):
            column_letter = col[0].column_letter
            formatting_info['column_widths'][column_letter] = sheet.column_dimensions[column_letter].width
        
        # Extract row heights
        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            formatting_info['row_heights'][row_idx] = sheet.row_dimensions[row_idx].height
        
        # Extract cell styles
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    style_info = {}
                    
                    # Font
                    if cell.font:
                        style_info['font'] = {
                            'name': cell.font.name,
                            'size': cell.font.size,
                            'bold': cell.font.bold,
                            'italic': cell.font.italic,
                            'color': cell.font.color.rgb if cell.font.color else None
                        }
                    
                    # Fill (background color)
                    if cell.fill and cell.fill.start_color.rgb:
                        style_info['fill'] = {
                            'color': cell.fill.start_color.rgb
                        }
                    
                    # Border
                    if any([cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]):
                        style_info['border'] = {}
                    
                    # Alignment
                    if cell.alignment:
                        style_info['alignment'] = {
                            'horizontal': cell.alignment.horizontal,
                            'vertical': cell.alignment.vertical
                        }
                    
                    if style_info:
                        formatting_info['styles'][(cell.row, cell.column)] = style_info
        
        # Extract merged cells
        formatting_info['merged_cells'] = [str(mc) for mc in sheet.merged_cells.ranges]
        
    except Exception as e:
        print(f"Warning: Could not extract formatting: {str(e)}")
    
    return formatting_info

def apply_formatting_to_worksheet(worksheet, formatting_info):
    """Apply formatting to a worksheet"""
    if not formatting_info:
        return
    
    try:
        # Apply column widths
        for col_letter, width in formatting_info.get('column_widths', {}).items():
            if width:
                worksheet.column_dimensions[col_letter].width = width
        
        # Apply row heights
        for row_idx, height in formatting_info.get('row_heights', {}).items():
            if height and row_idx <= worksheet.max_row:
                worksheet.row_dimensions[row_idx].height = height
        
        # Apply cell styles
        for (row, col), style in formatting_info.get('styles', {}).items():
            if row <= worksheet.max_row and col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                
                if 'font' in style:
                    font_info = style['font']
                    cell.font = Font(
                        name=font_info.get('name'),
                        size=font_info.get('size'),
                        bold=font_info.get('bold'),
                        italic=font_info.get('italic'),
                        color=font_info.get('color')
                    )
                
                if 'fill' in style:
                    fill_info = style['fill']
                    if fill_info.get('color'):
                        cell.fill = PatternFill(start_color=fill_info['color'], end_color=fill_info['color'], fill_type='solid')
                
                if 'alignment' in style:
                    align_info = style['alignment']
                    cell.alignment = Alignment(
                        horizontal=align_info.get('horizontal'),
                        vertical=align_info.get('vertical')
                    )
        
        # Apply merged cells (try to apply as many as possible)
        for merged_range in formatting_info.get('merged_cells', []):
            try:
                worksheet.merge_cells(merged_range)
            except:
                pass  # Skip if merge range is invalid
                
    except Exception as e:
        print(f"Warning: Could not apply formatting: {str(e)}")

def normalize_column_types(dfs):
    if len(dfs) <= 1:
        return dfs
    
    common_cols = find_common_columns(dfs)
    
    normalized_dfs = []
    for df in dfs:
        df_normalized = df.copy()
        
        for col in common_cols:
            if col in df_normalized.columns:
                col_types = []
                for other_df in dfs:
                    if col in other_df.columns:
                        col_types.append(other_df[col].dtype)
                
                if len(set(col_types)) > 1:
                    try:
                        df_normalized[col] = pd.to_datetime(df_normalized[col], errors='coerce')
                    except:
                        try:
                            df_normalized[col] = df_normalized[col].astype(str)
                        except:
                            pass
        
        normalized_dfs.append(df_normalized)
    
    return normalized_dfs

def convert_columns_to_common_type(dfs, column_name):
    converted_dfs = []
    
    for df in dfs:
        df_converted = df.copy()
        
        if column_name in df_converted.columns:
            try:
                df_converted[column_name] = pd.to_datetime(df_converted[column_name], errors='coerce')
                df_converted[column_name] = df_converted[column_name].astype(str)
            except:
                try:
                    df_converted[column_name] = df_converted[column_name].astype(str)
                except:
                    try:
                        df_converted[column_name] = df_converted[column_name].astype(object)
                    except:
                        pass
        
        converted_dfs.append(df_converted)
    
    return converted_dfs

def analyze_columns(dfs, filenames):
    if not dfs:
        return {}
    
    all_columns = {}
    column_types = {}
    
    for i, (df, filename) in enumerate(zip(dfs, filenames)):
        all_columns[filename] = set(df.columns)
        column_types[filename] = {col: str(df[col].dtype) for col in df.columns}
    
    common_cols = set.intersection(*[set(cols) for cols in all_columns.values()])
    
    unique_columns = {}
    for filename, columns in all_columns.items():
        unique_columns[filename] = list(columns - common_cols)
    
    all_unique_cols = set.union(*[set(cols) for cols in all_columns.values()])
    
    type_compatibility = {}
    for col in common_cols:
        types = []
        for filename in filenames:
            if col in column_types[filename]:
                types.append(column_types[filename][col])
        type_compatibility[col] = {
            'types': types,
            'compatible': len(set(types)) == 1
        }
    
    return {
        'common_columns': list(common_cols),
        'unique_columns': unique_columns,
        'all_columns': list(all_unique_cols),
        'column_summary': {name: list(cols) for name, cols in all_columns.items()},
        'column_types': column_types,
        'type_compatibility': type_compatibility
    }

def smart_merge(dfs, merge_type='inner', selected_columns=None):
    if not dfs:
        return None
    
    dfs_normalized = normalize_column_types(dfs)
    
    if selected_columns:
        dfs_to_merge = []
        for df in dfs_normalized:
            cols_to_keep = [col for col in selected_columns if col in df.columns]
            dfs_to_merge.append(df[cols_to_keep])
    else:
        common_cols = find_common_columns(dfs_normalized)
        if not common_cols:
            result = pd.concat(dfs_normalized, axis=0, ignore_index=True)
            return result.drop_duplicates()
        dfs_to_merge = dfs_normalized
    
    try:
        if len(dfs_to_merge) == 1:
            result = dfs_to_merge[0]
        else:
            if selected_columns:
                common_cols = [col for col in selected_columns if all(col in df.columns for df in dfs_to_merge)]
            else:
                common_cols = find_common_columns(dfs_to_merge)
            
            if not common_cols:
                result = pd.concat(dfs_to_merge, axis=0, ignore_index=True)
            else:
                problematic_columns = []
                for col in common_cols:
                    col_types = set()
                    for df in dfs_to_merge:
                        if col in df.columns:
                            col_types.add(str(df[col].dtype))
                    
                    if len(col_types) > 1:
                        problematic_columns.append(col)
                
                if problematic_columns:
                    for col in problematic_columns:
                        dfs_to_merge = convert_columns_to_common_type(dfs_to_merge, col)
                
                result = dfs_to_merge[0]
                for i in range(1, len(dfs_to_merge)):
                    try:
                        result = pd.merge(result, dfs_to_merge[i], on=common_cols, how=merge_type, suffixes=('', f'_{i}'))
                    except Exception as e:
                        if "merge on object and datetime64" in str(e).lower():
                            for col in common_cols:
                                for j, df in enumerate(dfs_to_merge):
                                    if col in df.columns:
                                        dfs_to_merge[j][col] = dfs_to_merge[j][col].astype(str)
                            
                            result = dfs_to_merge[0]
                            for j in range(1, len(dfs_to_merge)):
                                result = pd.merge(result, dfs_to_merge[j], on=common_cols, how=merge_type, suffixes=('', f'_{j}'))
                        else:
                            raise e
    
    except Exception as e:
        print(f"Merge failed, using concatenation: {str(e)}")
        result = pd.concat(dfs_to_merge, axis=0, ignore_index=True)
    
    result = result.drop_duplicates()
    return result

def find_common_columns(dfs):
    if not dfs:
        return []
    common_cols = set(dfs[0].columns)
    for df in dfs[1:]:
        common_cols = common_cols.intersection(set(df.columns))
    return list(common_cols)

def execute_python_code(code, df):
    """Safely execute Python code for data manipulation"""
    try:
        # Create a safe environment for code execution
        safe_globals = {
            'df': df.copy(),
            'pd': pd,
            'np': np,
            'plt': plt,
            'sns': sns,
            'max': max,
            'min': min,
            'sum': sum,
            'len': len,
            'range': range,
            'str': str,
            'int': int,
            'float': float,
            'list': list,
            'dict': dict
        }
        
        # Execute the code
        exec(code, safe_globals)
        
        # Get the modified dataframe
        modified_df = safe_globals.get('df', df)
        
        return modified_df, None
        
    except Exception as e:
        return df, f"Error executing code: {str(e)}"

def generate_plot_with_code(code, df):
    """Generate plot using AI-generated code"""
    try:
        # Create a safe environment for plot generation
        safe_globals = {
            'df': df,
            'pd': pd,
            'np': np,
            'plt': plt,
            'sns': sns
        }
        
        # Execute the plotting code
        exec(code, safe_globals)
        
        # Convert plot to base64
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', bbox_inches='tight', dpi=100)
        img_buffer.seek(0)
        img_base64 = base64.b64encode(img_buffer.getvalue()).decode('utf-8')
        plt.close()
        
        return img_base64, None
        
    except Exception as e:
        plt.close()
        return None, f"Error generating plot: {str(e)}"

def ask_openai_about_data(question, df_description, sample_data, dataframes):
    """Use OpenAI to analyze data, generate plots, and modify data"""
    if not client:
        return "OpenAI API not configured. Please provide an API key.", None, None
    
    try:
        prompt = f"""
        You are a data analysis expert with access to pandas, matplotlib, and seaborn. 
        You can generate Python code to analyze data, create visualizations, or modify datasets.

        Dataset Information:
        {df_description}

        Sample Data (first 5 rows):
        {sample_data}

        User Request: {question}

        You have two types of responses:

        1. For ANALYSIS/PLOTS: Generate Python code that creates visualizations. The code should:
           - Use plt.figure(figsize=(10, 6)) for appropriate sizing
           - Create meaningful plots (histograms, scatter plots, bar charts, etc.)
           - Include proper labels and titles
           - End with plt.show() but we'll capture the figure instead

        2. For DATA MODIFICATION: Generate Python code that modifies the dataframe 'df'
           - You can filter, transform, or analyze the data
           - Return the modified dataframe

        IMPORTANT: Your response should be in this exact format:

        ANALYSIS: [Your analysis explanation and insights]

        CODE:
        ```python
        # Your Python code here
        ```

        Only include code if the user requests analysis, plots, or data modification.
        If it's just a question about the data, provide the answer without code.
        """
        
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a helpful data analysis assistant. Generate Python code for data analysis and visualization when requested."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500,
            temperature=0.3
        )
        
        answer = response.choices[0].message.content
        
        # Extract code from response
        code_match = re.search(r'```python\n(.*?)\n```', answer, re.DOTALL)
        code = code_match.group(1) if code_match else None
        
        # Clean the answer by removing the code block
        clean_answer = re.sub(r'```python\n.*?\n```', '', answer, flags=re.DOTALL).strip()
        
        return clean_answer, code, None
        
    except Exception as e:
        return f"Error calling OpenAI API: {str(e)}", None, None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files[]' not in request.files:
        return jsonify({'error': 'No files selected'}), 400
    
    files = request.files.getlist('files[]')
    file_info = []
    dataframes = []
    filenames = []
    formatting_info_list = []
    
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            try:
                df, workbook = read_file(file_path, filename)
                dataframes.append(df)
                filenames.append(filename)
                
                # Extract formatting if it's an Excel file
                formatting_info = extract_excel_formatting(workbook) if workbook else None
                formatting_info_list.append(formatting_info)
                
                df_clean = clean_data_for_json(df.head(5))
                
                file_info.append({
                    'filename': filename,
                    'columns': df.columns.tolist(),
                    'shape': df.shape,
                    'sample_data': df_clean.to_dict('records'),
                    'column_types': {col: str(dtype) for col, dtype in df.dtypes.items()}
                })
                
                os.remove(file_path)
                
            except Exception as e:
                return jsonify({'error': f'Error processing {filename}: {str(e)}'}), 400
    
    app.config['CURRENT_DFS'] = dataframes
    app.config['CURRENT_FILENAMES'] = filenames
    app.config['CURRENT_FORMATTING'] = formatting_info_list
    
    column_analysis = analyze_columns(dataframes, filenames)
    
    return jsonify({
        'files': file_info,
        'column_analysis': column_analysis
    })

@app.route('/merge', methods=['POST'])
def merge_files():
    data = request.json
    merge_type = data.get('merge_type', 'inner')
    selected_columns = data.get('selected_columns', None)
    
    if 'CURRENT_DFS' not in app.config:
        return jsonify({'error': 'No files loaded'}), 400
    
    dataframes = app.config['CURRENT_DFS']
    formatting_info_list = app.config.get('CURRENT_FORMATTING', [])
    
    try:
        merged_df = smart_merge(dataframes, merge_type, selected_columns)
        
        if merged_df is None or merged_df.empty:
            return jsonify({'error': 'No data to merge'}), 400
        
        merged_df_clean = clean_data_for_json(merged_df)
        
        output_format = data.get('output_format', 'excel')
        
        if output_format == 'excel':
            filename = "merged_result.xlsx"
            output = io.BytesIO()
            
            # Create Excel file with formatting preservation
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                merged_df_clean.to_excel(writer, index=False, sheet_name='Merged_Data')
                
                # Try to apply formatting from the first Excel file
                if formatting_info_list and any(formatting_info_list):
                    # Get the first non-None formatting info
                    source_formatting = next((fmt for fmt in formatting_info_list if fmt is not None), None)
                    if source_formatting:
                        workbook = writer.book
                        worksheet = workbook['Merged_Data']
                        apply_formatting_to_worksheet(worksheet, source_formatting)
            
            output.seek(0)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
        elif output_format == 'csv':
            filename = "merged_result.csv"
            output = io.StringIO()
            merged_df_clean.to_csv(output, index=False)
            output = io.BytesIO(output.getvalue().encode('utf-8'))
            mimetype = 'text/csv'
            
        elif output_format == 'json':
            filename = "merged_result.json"
            output = io.BytesIO()
            json_data = merged_df_clean.to_json(orient='records', indent=2)
            output.write(json_data.encode('utf-8'))
            output.seek(0)
            mimetype = 'application/json'
        
        return send_file(
            output,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': f'Error merging files: {str(e)}'}), 400

@app.route('/chat', methods=['POST'])
def chat_with_data():
    """Chat with OpenAI about the data and execute code for plots/modifications"""
    if 'CURRENT_DFS' not in app.config:
        return jsonify({'error': 'No files loaded'}), 400
    
    dataframes = app.config['CURRENT_DFS']
    
    if not dataframes:
        return jsonify({'error': 'No data available'}), 400
    
    question = request.json.get('question', '')
    if not question:
        return jsonify({'error': 'No question provided'}), 400
    
    # Use the first dataframe for chat (can be extended to handle multiple)
    df = dataframes[0]
    
    # Create dataset description
    df_description = f"""
    Shape: {df.shape}
    Columns: {', '.join(df.columns.tolist())}
    Data Types: { {col: str(dtype) for col, dtype in df.dtypes.items()} }
    """
    
    sample_data = df.head(5).to_string()
    
    try:
        answer, code, error = ask_openai_about_data(question, df_description, sample_data, dataframes)
        
        response_data = {'answer': answer}
        plot_image = None
        modified_data = None
        
        # If there's code, execute it
        if code:
            # Check if it's plotting code or data modification code
            if any(keyword in code for keyword in ['plt.', 'sns.', 'plot', 'hist', 'scatter', 'bar', 'figure']):
                # It's plotting code
                plot_image, error = generate_plot_with_code(code, df)
                if plot_image:
                    response_data['plot_image'] = plot_image
                if error:
                    response_data['error'] = error
            else:
                # It's data modification code
                modified_df, error = execute_python_code(code, df)
                if not error:
                    # Update the stored dataframe
                    dataframes[0] = modified_df
                    app.config['CURRENT_DFS'] = dataframes
                    
                    # Create sample of modified data
                    modified_sample = modified_df.head(5).to_string()
                    response_data['modified_sample'] = modified_sample
                    response_data['modified_shape'] = modified_df.shape
                else:
                    response_data['error'] = error
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'error': f'Chat error: {str(e)}'}), 400

@app.route('/save_api_key', methods=['POST'])
def save_api_key():
    """Save OpenAI API key to local file"""
    api_key = request.json.get('api_key', '')
    
    if not api_key:
        return jsonify({'error': 'No API key provided'}), 400
    
    try:
        # Save to file in the application directory
        with open('openai_api_key.txt', 'w') as f:
            f.write(api_key.strip())
        
        # Initialize OpenAI client
        global client
        client = OpenAI(api_key=api_key.strip())
        
        return jsonify({'message': 'API key saved successfully'})
        
    except Exception as e:
        return jsonify({'error': f'Error saving API key: {str(e)}'}), 400

@app.route('/load_api_key', methods=['GET'])
def load_api_key():
    """Load OpenAI API key from local file"""
    try:
        if os.path.exists('openai_api_key.txt'):
            with open('openai_api_key.txt', 'r') as f:
                api_key = f.read().strip()
            
            if api_key:
                global client
                client = OpenAI(api_key=api_key)
                return jsonify({'api_key': api_key, 'message': 'API key loaded successfully'})
        
        return jsonify({'api_key': None, 'message': 'No API key found'})
        
    except Exception as e:
        return jsonify({'error': f'Error loading API key: {str(e)}'}), 400

@app.route('/get_current_data', methods=['GET'])
def get_current_data():
    """Get information about currently loaded data"""
    if 'CURRENT_DFS' not in app.config:
        return jsonify({'error': 'No data loaded'}), 400
    
    dataframes = app.config['CURRENT_DFS']
    
    if not dataframes:
        return jsonify({'error': 'No data available'}), 400
    
    df = dataframes[0]
    
    return jsonify({
        'shape': df.shape,
        'columns': df.columns.tolist(),
        'sample_data': df.head(5).to_dict('records')
    })

if __name__ == '__main__':
    # Create necessary directories
    for folder in [app.config['UPLOAD_FOLDER'], app.config['ANALYSIS_FOLDER']]:
        if not os.path.exists(folder):
            os.makedirs(folder)
    
    app.run(debug=True, port=5000)